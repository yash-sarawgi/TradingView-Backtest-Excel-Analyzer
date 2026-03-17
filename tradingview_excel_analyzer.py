#!/usr/bin/env python3
"""
TradingView Backtest Analyzer
Automatically adds comprehensive analysis to TradingView backtest exports,
then generates a professional PDF report from the same data.

Usage:
    python tradingview_analyzer.py <input_file.xlsx>
    python tradingview_analyzer.py <input_file.xlsx> <output_file.xlsx>

For batch processing:
    python tradingview_analyzer.py --batch <directory>

Outputs per file:
    <name>_analyzed.xlsx   — Excel workbook with Advanced Analysis sheet
    <name>_analyzed.pdf    — PDF backtest report
"""

""""
By - Yash Sarawgi, CFTe, CMT L3
"""

import sys
import os
import re
import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.dates as mdates

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable,
)


# ─────────────────────────────────────────────────────────────────────────────
# PDF COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
DARK_BG      = colors.HexColor("#0D1117")
CARD_BG      = colors.HexColor("#161B22")
ACCENT_BLUE  = colors.HexColor("#1F6FEB")
ACCENT_GREEN = colors.HexColor("#3FB950")
ACCENT_RED   = colors.HexColor("#F85149")
ACCENT_GOLD  = colors.HexColor("#D29922")
TEXT_PRIMARY = colors.HexColor("#E6EDF3")
TEXT_MUTED   = colors.HexColor("#8B949E")
BORDER       = colors.HexColor("#30363D")
ROW_EVEN     = colors.HexColor("#161B22")
ROW_ODD      = colors.HexColor("#1C2128")

PAGE_W, PAGE_H = A4
MARGIN = 18 * mm


# =============================================================================
# SECTION 1 — EXCEL ANALYSIS
# =============================================================================

def analyze_trades(trades_df):
    """Calculate all trading metrics from trades dataframe."""
    exit_trades = trades_df[trades_df['Type'].str.contains('Exit', case=False, na=False)].copy()

    if len(exit_trades) == 0:
        print("Warning: No exit trades found in data")
        return None, None

    exit_trades['Day of Week'] = exit_trades['Date and time'].dt.day_name()
    exit_trades['Hour']        = exit_trades['Date and time'].dt.hour
    exit_trades['Month']       = exit_trades['Date and time'].dt.month_name()
    exit_trades['Year']        = exit_trades['Date and time'].dt.year
    exit_trades['Date']        = exit_trades['Date and time'].dt.date

    total_trades    = len(exit_trades)
    winning_trades  = len(exit_trades[exit_trades['Net P&L USD'] > 0])
    losing_trades   = len(exit_trades[exit_trades['Net P&L USD'] < 0])
    win_rate        = (winning_trades / total_trades * 100) if total_trades > 0 else 0

    total_profit    = exit_trades[exit_trades['Net P&L USD'] > 0]['Net P&L USD'].sum()
    total_loss      = abs(exit_trades[exit_trades['Net P&L USD'] < 0]['Net P&L USD'].sum())
    profit_factor   = (total_profit / total_loss) if total_loss > 0 else 0

    avg_win   = exit_trades[exit_trades['Net P&L USD'] > 0]['Net P&L USD'].mean() if winning_trades > 0 else 0
    avg_loss  = exit_trades[exit_trades['Net P&L USD'] < 0]['Net P&L USD'].mean() if losing_trades  > 0 else 0
    avg_trade = exit_trades['Net P&L USD'].mean()

    largest_win  = exit_trades['Net P&L USD'].max()
    largest_loss = exit_trades['Net P&L USD'].min()

    max_consecutive_wins = max_consecutive_losses = 0
    cur_wins = cur_losses = 0
    for pnl in exit_trades['Net P&L USD']:
        if pnl > 0:
            cur_wins += 1; cur_losses = 0
            max_consecutive_wins = max(max_consecutive_wins, cur_wins)
        else:
            cur_losses += 1; cur_wins = 0
            max_consecutive_losses = max(max_consecutive_losses, cur_losses)

    dow_order  = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    dow_stats  = exit_trades.groupby('Day of Week').agg({'Net P&L USD': ['sum', 'count', 'mean']}).round(2)
    dow_stats.columns = ['Total P&L', 'Trades', 'Avg P&L']
    dow_stats['Win Rate %'] = exit_trades.groupby('Day of Week').apply(
        lambda x: (len(x[x['Net P&L USD'] > 0]) / len(x) * 100) if len(x) > 0 else 0
    ).round(2)
    dow_stats = dow_stats.reindex([d for d in dow_order if d in dow_stats.index])

    hour_stats  = exit_trades.groupby('Hour').agg({'Net P&L USD': ['sum', 'count', 'mean']}).round(2)
    hour_stats.columns = ['Total P&L', 'Trades', 'Avg P&L']
    hour_stats['Win Rate %'] = exit_trades.groupby('Hour').apply(
        lambda x: (len(x[x['Net P&L USD'] > 0]) / len(x) * 100) if len(x) > 0 else 0
    ).round(2)

    month_stats  = exit_trades.groupby('Month').agg({'Net P&L USD': ['sum', 'count', 'mean']}).round(2)
    month_stats.columns = ['Total P&L', 'Trades', 'Avg P&L']
    month_stats['Win Rate %'] = exit_trades.groupby('Month').apply(
        lambda x: (len(x[x['Net P&L USD'] > 0]) / len(x) * 100) if len(x) > 0 else 0
    ).round(2)

    daily_pnl = exit_trades.groupby('Date')['Net P&L USD'].sum().reset_index()
    daily_pnl['Cumulative P&L'] = daily_pnl['Net P&L USD'].cumsum()

    exit_trades_sorted = exit_trades.sort_values('Date and time')
    cumulative_pnl = exit_trades_sorted['Net P&L USD'].cumsum()
    running_max    = cumulative_pnl.expanding().max()
    drawdown       = cumulative_pnl - running_max
    max_drawdown   = drawdown.min()
    max_drawdown_pct = (max_drawdown / running_max[drawdown.idxmin()]) * 100 \
                       if running_max[drawdown.idxmin()] != 0 else 0

    returns    = exit_trades['Net P&L %'].values
    avg_return = returns.mean()
    std_return = returns.std()
    sharpe_like = (avg_return / std_return) if std_return != 0 else 0

    expectancy = (
        exit_trades[exit_trades['Net P&L USD'] > 0]['Net P&L USD'].mean() *
        (exit_trades[exit_trades['Net P&L USD'] > 0].shape[0] / exit_trades.shape[0])
    ) + (
        exit_trades[exit_trades['Net P&L USD'] < 0]['Net P&L USD'].mean() *
        (exit_trades[exit_trades['Net P&L USD'] < 0].shape[0] / exit_trades.shape[0])
    )

    distribution = [
        ('< -$500',       len(exit_trades[exit_trades['Net P&L USD'] < -500])),
        ('-$500 to -$250', len(exit_trades[(exit_trades['Net P&L USD'] >= -500) & (exit_trades['Net P&L USD'] < -250)])),
        ('-$250 to -$100', len(exit_trades[(exit_trades['Net P&L USD'] >= -250) & (exit_trades['Net P&L USD'] < -100)])),
        ('-$100 to $0',    len(exit_trades[(exit_trades['Net P&L USD'] >= -100) & (exit_trades['Net P&L USD'] <    0)])),
        ('$0 to $100',     len(exit_trades[(exit_trades['Net P&L USD'] >=    0) & (exit_trades['Net P&L USD'] <  100)])),
        ('$100 to $250',   len(exit_trades[(exit_trades['Net P&L USD'] >=  100) & (exit_trades['Net P&L USD'] <  250)])),
        ('$250 to $500',   len(exit_trades[(exit_trades['Net P&L USD'] >=  250) & (exit_trades['Net P&L USD'] <  500)])),
        ('> $500',         len(exit_trades[exit_trades['Net P&L USD'] >= 500])),
    ]

    metrics = {
        'total_trades':           int(total_trades),
        'winning_trades':         int(winning_trades),
        'losing_trades':          int(losing_trades),
        'win_rate':               float(win_rate),
        'total_profit':           float(total_profit),
        'total_loss':             float(total_loss),
        'profit_factor':          float(profit_factor),
        'avg_win':                float(avg_win),
        'avg_loss':               float(avg_loss),
        'avg_trade':              float(avg_trade),
        'largest_win':            float(largest_win),
        'largest_loss':           float(largest_loss),
        'max_consecutive_wins':   int(max_consecutive_wins),
        'max_consecutive_losses': int(max_consecutive_losses),
        'max_drawdown':           float(max_drawdown),
        'max_drawdown_pct':       float(max_drawdown_pct),
        'sharpe_like':            float(sharpe_like),
        'expectancy':             float(expectancy),
        'std_return':             float(std_return),
        'avg_return':             float(avg_return),
    }

    data_frames = {
        'dow_stats':    dow_stats,
        'hour_stats':   hour_stats,
        'month_stats':  month_stats,
        'daily_pnl':    daily_pnl,
        'distribution': distribution,
    }

    return metrics, data_frames


def create_analysis_sheet(wb, metrics, data_frames):
    """Create the Advanced Analysis sheet in the workbook."""
    if 'Advanced Analysis' in wb.sheetnames:
        del wb['Advanced Analysis']
    ws = wb.create_sheet('Advanced Analysis', 0)

    header_font    = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill    = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_font = Font(name='Arial', size=11, bold=True)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    metric_font    = Font(name='Arial', size=10)

    ws['A1'] = 'TradingView Backtest - Advanced Analysis'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # KEY METRICS
    row = 3
    ws[f'A{row}'] = 'KEY PERFORMANCE METRICS'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    metrics_data = [
        ('Total Trades',           metrics['total_trades'],                      ''),
        ('Winning Trades',         metrics['winning_trades'],                     ''),
        ('Losing Trades',          metrics['losing_trades'],                      ''),
        ('Win Rate',               metrics['win_rate'],                           '%'),
        ('',                       '',                                            ''),
        ('Total Profit',           metrics['total_profit'],                       '$'),
        ('Total Loss',             abs(metrics['total_loss']),                    '$'),
        ('Net Profit',             metrics['total_profit'] - abs(metrics['total_loss']), '$'),
        ('Profit Factor',          metrics['profit_factor'],                      ''),
        ('',                       '',                                            ''),
        ('Average Win',            metrics['avg_win'],                            '$'),
        ('Average Loss',           abs(metrics['avg_loss']),                      '$'),
        ('Average Trade',          metrics['avg_trade'],                          '$'),
        ('',                       '',                                            ''),
        ('Largest Win',            metrics['largest_win'],                        '$'),
        ('Largest Loss',           abs(metrics['largest_loss']),                  '$'),
        ('',                       '',                                            ''),
        ('Max Consecutive Wins',   metrics['max_consecutive_wins'],               ''),
        ('Max Consecutive Losses', metrics['max_consecutive_losses'],             ''),
    ]

    for metric_name, value, suffix in metrics_data:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = metric_font
        if metric_name:
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
            if suffix == '$':
                ws[f'B{row}'].number_format = '$#,##0.00'
            elif suffix == '%':
                ws[f'B{row}'].number_format = '0.00"%"'
            else:
                ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

    # DAY OF WEEK
    ws[f'D{3}'] = 'DAY OF WEEK ANALYSIS'
    ws[f'D{3}'].font = header_font
    ws[f'D{3}'].fill = header_fill
    ws.merge_cells(f'D{3}:G{3}')

    row = 4
    for col_idx, header in enumerate(['Day', 'Total P&L', 'Trades', 'Avg P&L', 'Win Rate %'], start=4):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font  = subheader_font
        cell.fill  = subheader_fill
        cell.alignment = Alignment(horizontal='center')

    row += 1
    dow_start_row = row
    for idx, data_row in data_frames['dow_stats'].iterrows():
        ws.cell(row=row, column=4, value=idx)
        ws.cell(row=row, column=5, value=data_row['Total P&L']).number_format  = '$#,##0.00'
        ws.cell(row=row, column=6, value=data_row['Trades']).number_format     = '#,##0'
        ws.cell(row=row, column=7, value=data_row['Avg P&L']).number_format    = '$#,##0.00'
        ws.cell(row=row, column=8, value=data_row['Win Rate %']).number_format = '0.00"%"'
        row += 1
    dow_end_row = row - 1

    # HOUR ANALYSIS
    row += 1
    ws[f'D{row}'] = 'HOUR OF DAY ANALYSIS'
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].fill = header_fill
    ws.merge_cells(f'D{row}:G{row}')

    row += 1
    for col_idx, header in enumerate(['Hour', 'Total P&L', 'Trades', 'Avg P&L', 'Win Rate %'], start=4):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font  = subheader_font
        cell.fill  = subheader_fill
        cell.alignment = Alignment(horizontal='center')

    row += 1
    hour_start_row = row
    for idx, data_row in data_frames['hour_stats'].iterrows():
        ws.cell(row=row, column=4, value=int(idx))
        ws.cell(row=row, column=5, value=data_row['Total P&L']).number_format  = '$#,##0.00'
        ws.cell(row=row, column=6, value=data_row['Trades']).number_format     = '#,##0'
        ws.cell(row=row, column=7, value=data_row['Avg P&L']).number_format    = '$#,##0.00'
        ws.cell(row=row, column=8, value=data_row['Win Rate %']).number_format = '0.00"%"'
        row += 1
    hour_end_row = row - 1

    # MONTH ANALYSIS
    row += 1
    ws[f'D{row}'] = 'MONTH ANALYSIS'
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].fill = header_fill
    ws.merge_cells(f'D{row}:G{row}')

    row += 1
    for col_idx, header in enumerate(['Month', 'Total P&L', 'Trades', 'Avg P&L', 'Win Rate %'], start=4):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font  = subheader_font
        cell.fill  = subheader_fill
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for idx, data_row in data_frames['month_stats'].iterrows():
        ws.cell(row=row, column=4, value=idx)
        ws.cell(row=row, column=5, value=data_row['Total P&L']).number_format  = '$#,##0.00'
        ws.cell(row=row, column=6, value=data_row['Trades']).number_format     = '#,##0'
        ws.cell(row=row, column=7, value=data_row['Avg P&L']).number_format    = '$#,##0.00'
        ws.cell(row=row, column=8, value=data_row['Win Rate %']).number_format = '0.00"%"'
        row += 1

    # TRADE P&L DISTRIBUTION
    dist_row = 27
    ws[f'J{dist_row}'] = 'TRADE P&L DISTRIBUTION'
    ws[f'J{dist_row}'].font = header_font
    ws[f'J{dist_row}'].fill = header_fill
    ws.merge_cells(f'J{dist_row}:K{dist_row}')

    dist_row += 1
    ws.cell(row=dist_row, column=10, value='P&L Range').font = subheader_font
    ws.cell(row=dist_row, column=11, value='Count').font     = subheader_font
    ws.cell(row=dist_row, column=12, value='%').font         = subheader_font

    dist_row += 1
    for range_label, count in data_frames['distribution']:
        ws.cell(row=dist_row, column=10, value=range_label)
        ws.cell(row=dist_row, column=11, value=count)
        ws.cell(row=dist_row, column=12,
                value=count / metrics['total_trades'] * 100).number_format = '0.00"%"'
        dist_row += 1

    # DRAWDOWN & RISK METRICS
    risk_row = dist_row + 2
    ws[f'J{risk_row}'] = 'DRAWDOWN & RISK METRICS'
    ws[f'J{risk_row}'].font = header_font
    ws[f'J{risk_row}'].fill = header_fill
    ws.merge_cells(f'J{risk_row}:K{risk_row}')

    risk_row += 1
    risk_metrics_data = [
        ('Max Drawdown ($)',      abs(metrics['max_drawdown']),     '$'),
        ('Max Drawdown (%)',      abs(metrics['max_drawdown_pct']), '%'),
        ('Sharpe-like Ratio',     metrics['sharpe_like'],           ''),
        ('Expectancy per Trade',  metrics['expectancy'],            '$'),
        ('Std Dev of Returns',    metrics['std_return'],            '%'),
        ('Avg Return per Trade',  metrics['avg_return'],            '%'),
    ]
    for metric_name, value, suffix in risk_metrics_data:
        ws.cell(row=risk_row, column=10, value=metric_name)
        ws.cell(row=risk_row, column=11, value=value)
        if suffix == '$':
            ws.cell(row=risk_row, column=11).number_format = '$#,##0.00'
        elif suffix == '%':
            ws.cell(row=risk_row, column=11).number_format = '0.00"%"'
        else:
            ws.cell(row=risk_row, column=11).number_format = '#,##0.00'
        risk_row += 1

    # DAILY EQUITY CURVE
    equity_row = row + 2
    ws[f'A{equity_row}'] = 'DAILY EQUITY CURVE'
    ws[f'A{equity_row}'].font = header_font
    ws[f'A{equity_row}'].fill = header_fill
    ws.merge_cells(f'A{equity_row}:C{equity_row}')

    equity_row += 1
    ws.cell(row=equity_row, column=1, value='Date').font           = subheader_font
    ws.cell(row=equity_row, column=1).fill                         = subheader_fill
    ws.cell(row=equity_row, column=2, value='Daily P&L').font      = subheader_font
    ws.cell(row=equity_row, column=2).fill                         = subheader_fill
    ws.cell(row=equity_row, column=3, value='Cumulative P&L').font = subheader_font
    ws.cell(row=equity_row, column=3).fill                         = subheader_fill

    equity_row += 1
    equity_start_row = equity_row
    for _, data_row in data_frames['daily_pnl'].iterrows():
        ws.cell(row=equity_row, column=1, value=str(data_row['Date']))
        ws.cell(row=equity_row, column=2, value=data_row['Net P&L USD']).number_format  = '$#,##0.00'
        ws.cell(row=equity_row, column=3, value=data_row['Cumulative P&L']).number_format = '$#,##0.00'
        equity_row += 1
    equity_end_row = equity_row - 1

    # CHARTS
    chart1 = BarChart()
    chart1.type = "col"; chart1.style = 10
    chart1.title = "P&L by Day of Week"
    chart1.y_axis.title = 'Total P&L ($)'; chart1.x_axis.title = 'Day of Week'
    data = Reference(ws, min_col=5, min_row=4,          max_row=dow_end_row,  max_col=5)
    cats = Reference(ws, min_col=4, min_row=5,          max_row=dow_end_row)
    chart1.add_data(data, titles_from_data=True); chart1.set_categories(cats)
    chart1.height = 10; chart1.width = 20
    ws.add_chart(chart1, "J3")

    chart2 = BarChart()
    chart2.type = "col"; chart2.style = 11
    chart2.title = "P&L by Hour of Day"
    chart2.y_axis.title = 'Total P&L ($)'; chart2.x_axis.title = 'Hour'
    data = Reference(ws, min_col=5, min_row=hour_start_row - 1, max_row=hour_end_row, max_col=5)
    cats = Reference(ws, min_col=4, min_row=hour_start_row,     max_row=hour_end_row)
    chart2.add_data(data, titles_from_data=True); chart2.set_categories(cats)
    chart2.height = 10; chart2.width = 20
    ws.add_chart(chart2, "J20")

    chart3 = BarChart()
    chart3.type = "col"; chart3.style = 12
    chart3.title = "Win Rate by Day of Week"
    chart3.y_axis.title = 'Win Rate (%)'; chart3.x_axis.title = 'Day of Week'
    data = Reference(ws, min_col=8, min_row=4,          max_row=dow_end_row, max_col=8)
    cats = Reference(ws, min_col=4, min_row=5,          max_row=dow_end_row)
    chart3.add_data(data, titles_from_data=True); chart3.set_categories(cats)
    chart3.height = 10; chart3.width = 20
    ws.add_chart(chart3, "T3")

    chart4 = LineChart()
    chart4.style = 13
    chart4.title = "Cumulative P&L Equity Curve"
    chart4.y_axis.title = 'Cumulative P&L ($)'; chart4.x_axis.title = 'Trading Days'
    data = Reference(ws, min_col=3, min_row=equity_start_row - 1,
                     max_row=min(equity_end_row, equity_start_row + 200), max_col=3)
    chart4.add_data(data, titles_from_data=True)
    chart4.height = 12; chart4.width = 25
    ws.add_chart(chart4, "J37")

    for col, width in [('A', 25), ('B', 15), ('C', 15), ('D', 15), ('E', 15),
                       ('F', 12), ('G', 12), ('H', 12), ('J', 25), ('K', 15), ('L', 12)]:
        ws.column_dimensions[col].width = width


# =============================================================================
# SECTION 2 — PDF REPORT GENERATION
# =============================================================================

def _extract_strategy_name(filepath):
    base = os.path.splitext(os.path.basename(filepath))[0]
    base = re.sub(r"_analyzed$", "", base)
    base = re.sub(r"_\d{4}-\d{2}-\d{2}$", "", base)
    base = base.replace("___", " — ").replace("__", " ").replace("_", " ")
    return re.sub(r"\s+", " ", base).strip()


def _read_properties(wb):
    props = {}
    if "Properties" in wb.sheetnames:
        for row in wb["Properties"].iter_rows(values_only=True):
            if row[0] and row[1]:
                props[str(row[0]).strip()] = row[1]
    return props


def build_pdf_data(input_file, wb, metrics, data_frames):
    """
    Assemble the data dict needed by build_pdf() directly from
    already-computed metrics and data_frames — no disk re-read needed.
    """
    props = _read_properties(wb)

    stock      = str(props.get("Symbol", "Unknown")).strip()
    strategy   = _extract_strategy_name(input_file)
    date_range = str(props.get("Backtesting range",
                               props.get("Trading range", "N/A"))).strip()

    kpm = {
        "Total Trades":         metrics["total_trades"],
        "Winning Trades":       metrics["winning_trades"],
        "Losing Trades":        metrics["losing_trades"],
        "Win Rate (%)":         metrics["win_rate"],
        "Total Profit":         metrics["total_profit"],
        "Total Loss":           abs(metrics["total_loss"]),
        "Net Profit":           metrics["total_profit"] - abs(metrics["total_loss"]),
        "Profit Factor":        metrics["profit_factor"],
        "Average Win":          metrics["avg_win"],
        "Average Loss":         abs(metrics["avg_loss"]),
        "Average Trade":        metrics["avg_trade"],
        "Largest Win":          metrics["largest_win"],
        "Largest Loss":         abs(metrics["largest_loss"]),
        "Max Consec. Wins":     metrics["max_consecutive_wins"],
        "Max Consec. Losses":   metrics["max_consecutive_losses"],
    }

    dow_rows = [
        [idx,
         row["Total P&L"], row["Trades"],
         row["Avg P&L"],   row["Win Rate %"]]
        for idx, row in data_frames["dow_stats"].iterrows()
    ]

    hour_rows = [
        [idx,
         row["Total P&L"], row["Trades"],
         row["Avg P&L"],   row["Win Rate %"]]
        for idx, row in data_frames["hour_stats"].iterrows()
    ]

    month_rows = [
        [idx,
         row["Total P&L"], row["Trades"],
         row["Avg P&L"],   row["Win Rate %"]]
        for idx, row in data_frames["month_stats"].iterrows()
    ]

    total = metrics["total_trades"]
    dist_rows = [
        [label, count, (count / total * 100) if total > 0 else 0]
        for label, count in data_frames["distribution"]
    ]

    risk = {
        "Max Drawdown ($)":       abs(metrics["max_drawdown"]),
        "Max Drawdown (%)":       abs(metrics["max_drawdown_pct"]),
        "Sharpe-like Ratio":      metrics["sharpe_like"],
        "Expectancy per Trade":   metrics["expectancy"],
        "Std Dev of Returns":     metrics["std_return"],
        "Avg Return per Trade":   metrics["avg_return"],
    }

    dp = data_frames["daily_pnl"]
    dates     = [str(d) for d in dp["Date"].tolist()]
    daily_pnl = dp["Net P&L USD"].tolist()
    cum_pnl   = dp["Cumulative P&L"].tolist()

    return {
        "stock":          stock,
        "strategy":       strategy,
        "date_range":     date_range,
        "kpm":            kpm,
        "dow":            dow_rows,
        "hour":           hour_rows,
        "month":          month_rows,
        "dist":           dist_rows,
        "risk":           risk,
        "equity_dates":   dates,
        "equity_daily":   daily_pnl,
        "equity_cumul":   cum_pnl,
    }


# ── PDF helpers ───────────────────────────────────────────────────────────────

def _fmt(val, decimals=2, prefix="", suffix=""):
    if val is None:
        return "—"
    try:
        v = float(val)
        s = f"{prefix}{v:,.{decimals}f}{suffix}" if decimals else f"{prefix}{v:,.0f}{suffix}"
        return s
    except Exception:
        return str(val)


def _fig_to_image(fig, w, h):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return Image(buf, width=w, height=h)


def _make_styled_table(rows, col_widths, pnl_col=None):
    style = [
        ("BACKGROUND",   (0, 0), (-1,  0), CARD_BG),
        ("TEXTCOLOR",    (0, 0), (-1, -1), TEXT_PRIMARY),
        ("FONTNAME",     (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("FONTNAME",     (0, 0), (-1,  0), "Helvetica-Bold"),
        ("TEXTCOLOR",    (0, 0), (-1,  0), TEXT_MUTED),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [ROW_ODD, ROW_EVEN]),
        ("GRID",         (0, 0), (-1, -1), 0.3, BORDER),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]
    if pnl_col is not None:
        for i, row in enumerate(rows[1:], 1):
            try:
                v = float(str(row[pnl_col]).replace("$", "").replace(",", ""))
                clr = ACCENT_GREEN if v >= 0 else ACCENT_RED
                style.append(("TEXTCOLOR", (pnl_col, i), (pnl_col, i), clr))
            except Exception:
                pass
    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle(style))
    return t


def _chart_equity(data, w, h):
    dates, cum, daily = data["equity_dates"], data["equity_cumul"], data["equity_daily"]
    if not dates:
        return None

    parsed = []
    for d in dates:
        try:
            parsed.append(datetime.strptime(str(d)[:10], "%Y-%m-%d"))
        except Exception:
            parsed.append(d)

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(w / 72, h / 72),
                                   gridspec_kw={"height_ratios": [3, 1]},
                                   facecolor="#0D1117")
    fig.subplots_adjust(hspace=0.08)

    ax1.set_facecolor("#0D1117")
    ax1.fill_between(parsed, cum, 0, where=[v >= 0 for v in cum], color="#3FB950", alpha=0.15)
    ax1.fill_between(parsed, cum, 0, where=[v <  0 for v in cum], color="#F85149", alpha=0.15)
    ax1.plot(parsed, cum, color="#1F6FEB", linewidth=1.6, zorder=3)
    ax1.axhline(0, color="#30363D", linewidth=0.8, linestyle="--")

    min_v, max_v = min(cum), max(cum)
    min_i, max_i = cum.index(min_v), cum.index(max_v)
    ax1.scatter([parsed[min_i]], [min_v], color="#F85149", s=40, zorder=5)
    ax1.scatter([parsed[max_i]], [max_v], color="#3FB950", s=40, zorder=5)
    ax1.annotate(f"${min_v:+.1f}", (parsed[min_i], min_v),
                 textcoords="offset points", xytext=(6, -12), color="#F85149", fontsize=7)
    ax1.annotate(f"${max_v:+.1f}", (parsed[max_i], max_v),
                 textcoords="offset points", xytext=(6, 6),  color="#3FB950", fontsize=7)

    ax1.set_ylabel("Cumulative P&L ($)", color="#8B949E", fontsize=8)
    ax1.tick_params(colors="#8B949E", labelsize=7)
    ax1.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax1.set_xticklabels([])
    ax1.set_title("Daily Equity Curve", color="#E6EDF3", fontsize=9, pad=6, loc="left")
    for sp in ax1.spines.values():
        sp.set_color("#30363D")
    ax1.grid(axis="y", color="#30363D", linewidth=0.5, linestyle=":")

    ax2.set_facecolor("#0D1117")
    ax2.bar(parsed, daily, color=["#3FB950" if v >= 0 else "#F85149" for v in daily], width=0.8)
    ax2.axhline(0, color="#30363D", linewidth=0.8)
    ax2.set_ylabel("Daily P&L ($)", color="#8B949E", fontsize=7)
    ax2.tick_params(colors="#8B949E", labelsize=6.5)
    ax2.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
    ax2.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
    plt.setp(ax2.get_xticklabels(), rotation=30, ha="right")
    for sp in ax2.spines.values():
        sp.set_color("#30363D")
    ax2.grid(axis="y", color="#30363D", linewidth=0.4, linestyle=":")

    return _fig_to_image(fig, w, h)


def _chart_dow(dow_rows, w, h):
    if not dow_rows:
        return None
    days = [r[0] for r in dow_rows]
    vals = [float(r[1]) if r[1] is not None else 0 for r in dow_rows]

    fig, ax = plt.subplots(figsize=(w / 72, h / 72), facecolor="#0D1117")
    ax.set_facecolor("#161B22")
    bars = ax.bar(days, vals, color=["#3FB950" if v >= 0 else "#F85149" for v in vals],
                  width=0.55, edgecolor="#30363D", linewidth=0.6)
    ax.axhline(0, color="#30363D", linewidth=0.8)
    for bar, val in zip(bars, vals):
        yoff = max(vals) * 0.03 if val >= 0 else min(vals) * 0.03
        ax.text(bar.get_x() + bar.get_width() / 2, val + yoff,
                f"${val:+.0f}", ha="center",
                va="bottom" if val >= 0 else "top",
                color="#E6EDF3", fontsize=7.5, fontweight="bold")
    ax.set_title("Total P&L by Day of Week", color="#E6EDF3", fontsize=9, pad=6, loc="left")
    ax.tick_params(colors="#8B949E", labelsize=8)
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    for sp in ax.spines.values():
        sp.set_color("#30363D")
    ax.grid(axis="y", color="#30363D", linewidth=0.4, linestyle=":")
    fig.tight_layout()
    return _fig_to_image(fig, w, h)


def _chart_hour(hour_rows, w, h):
    if not hour_rows:
        return None
    hours = [f"{int(r[0]):02d}:00" if r[0] is not None else "?" for r in hour_rows]
    pnl   = [float(r[1]) if r[1] is not None else 0 for r in hour_rows]
    wr    = [float(r[4]) if r[4] is not None else 0 for r in hour_rows]

    fig, ax1 = plt.subplots(figsize=(w / 72, h / 72), facecolor="#0D1117")
    ax1.set_facecolor("#161B22")
    ax2 = ax1.twinx()

    x = range(len(hours))
    ax1.bar(x, pnl, color=["#3FB950" if v >= 0 else "#F85149" for v in pnl],
            width=0.55, edgecolor="#30363D", linewidth=0.6)
    ax2.plot(x, wr, color="#D29922", linewidth=1.5, marker="o", markersize=4)
    ax2.axhline(50, color="#30363D", linewidth=0.8, linestyle="--")

    ax1.set_xticks(list(x))
    ax1.set_xticklabels(hours, fontsize=7.5, color="#8B949E")
    ax1.tick_params(axis="y", colors="#8B949E", labelsize=7.5)
    ax2.tick_params(axis="y", colors="#D29922", labelsize=7.5)
    ax1.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"${v:,.0f}"))
    ax1.set_ylabel("Total P&L ($)", color="#8B949E", fontsize=7.5)
    ax2.set_ylabel("Win Rate (%)",  color="#D29922", fontsize=7.5)
    ax1.set_title("P&L & Win Rate by Hour", color="#E6EDF3", fontsize=9, pad=6, loc="left")
    for sp in ax1.spines.values():
        sp.set_color("#30363D")
    for sp in ax2.spines.values():
        sp.set_color("#30363D")
    ax1.grid(axis="y", color="#30363D", linewidth=0.4, linestyle=":")
    fig.tight_layout()
    return _fig_to_image(fig, w, h)


def _chart_dist(dist_rows, w, h):
    if not dist_rows:
        return None
    labels = [str(r[0]) for r in dist_rows]
    counts = [float(r[1]) if r[1] is not None else 0 for r in dist_rows]
    clrs   = ["#3FB950" if "$0" in l or ("to $" in l and "-" not in l) else "#F85149"
              for l in labels]

    fig, ax = plt.subplots(figsize=(w / 72, h / 72), facecolor="#0D1117")
    ax.set_facecolor("#161B22")
    bars = ax.barh(labels, counts, color=clrs, edgecolor="#30363D", linewidth=0.6, height=0.6)
    for bar, val in zip(bars, counts):
        if val > 0:
            ax.text(val + max(counts) * 0.01, bar.get_y() + bar.get_height() / 2,
                    f"{val:.0f}", va="center", color="#E6EDF3", fontsize=7.5)
    ax.set_title("Trade P&L Distribution", color="#E6EDF3", fontsize=9, pad=6, loc="left")
    ax.tick_params(colors="#8B949E", labelsize=7.5)
    ax.set_xlabel("Number of Trades", color="#8B949E", fontsize=7.5)
    for sp in ax.spines.values():
        sp.set_color("#30363D")
    ax.grid(axis="x", color="#30363D", linewidth=0.4, linestyle=":")
    fig.tight_layout()
    return _fig_to_image(fig, w, h)


def build_pdf(data, output_path):
    """Render the full PDF report from the assembled data dict."""
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=MARGIN,
    )
    cw = PAGE_W - 2 * MARGIN
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s   = ps("T",  fontSize=18, textColor=TEXT_PRIMARY, alignment=TA_CENTER,
                   fontName="Helvetica-Bold", spaceAfter=2)
    section_s = ps("S",  fontSize=11, textColor=ACCENT_BLUE,
                   fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    label_s   = ps("L",  fontSize=8,  textColor=TEXT_MUTED,  fontName="Helvetica")
    value_s   = ps("V",  fontSize=8,  textColor=TEXT_PRIMARY, fontName="Helvetica")

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("TradingView Backtest Report", title_s))
    story.append(Spacer(1, 2 * mm))

    hdr_data = [
        [Paragraph("<b>Stock / Symbol</b>", label_s),
         Paragraph(f"<font color='#1F6FEB'><b>{data['stock']}</b></font>", value_s),
         Paragraph("<b>Date Range</b>", label_s),
         Paragraph(data["date_range"], value_s)],
        [Paragraph("<b>Strategy</b>", label_s),
         Paragraph(data["strategy"], value_s),
         Paragraph("<b>Generated</b>", label_s),
         Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M"), value_s)],
    ]
    hdr_tbl = Table(hdr_data, colWidths=[cw * f for f in [0.15, 0.35, 0.15, 0.35]])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), CARD_BG),
        ("GRID",         (0, 0), (-1, -1), 0.4, BORDER),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT_BLUE,
                             spaceAfter=8, spaceBefore=6))

    # ── KPI cards ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Key Performance Metrics", section_s))
    kpm = data["kpm"]

    def _nc(val, green_if_positive=True):
        try:
            return "#3FB950" if (float(val) >= 0) == green_if_positive else "#F85149"
        except Exception:
            return "#E6EDF3"

    kpi_items = [
        ("Net Profit",    _fmt(kpm["Net Profit"], 2, "$"),       _nc(kpm["Net Profit"])),
        ("Profit Factor", _fmt(kpm["Profit Factor"], 3),
         "#3FB950" if float(kpm.get("Profit Factor", 0) or 0) >= 1 else "#F85149"),
        ("Win Rate",      _fmt(kpm["Win Rate (%)"], 2, suffix="%"),
         "#3FB950" if float(kpm.get("Win Rate (%)", 0) or 0) >= 50 else "#F85149"),
        ("Total Trades",  _fmt(kpm["Total Trades"], 0),            "#E6EDF3"),
        ("Avg Win",       _fmt(kpm["Average Win"],  2, "$"),        "#3FB950"),
        ("Avg Loss",      _fmt(kpm["Average Loss"], 2, "$"),        "#F85149"),
    ]
    kpi_cells = [
        Paragraph(
            f"<para align='center'>"
            f"<font color='#8B949E' size='7'>{lbl}</font><br/>"
            f"<font color='{clr}' size='13'><b>{val}</b></font></para>",
            styles["Normal"]
        )
        for lbl, val, clr in kpi_items
    ]
    kpi_tbl = Table([kpi_cells], colWidths=[cw / 6] * 6)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), CARD_BG),
        ("GRID",         (0, 0), (-1, -1), 0.4, BORDER),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 4 * mm))

    # Full KPM two-column table
    kpm_pairs = list(kpm.items())
    half = (len(kpm_pairs) + 1) // 2
    left_kpm, right_kpm = kpm_pairs[:half], kpm_pairs[half:]

    def _kpm_row(label, val):
        try:
            fv = float(val)
            dec = 0 if label in ("Total Trades", "Winning Trades", "Losing Trades",
                                  "Max Consec. Wins", "Max Consec. Losses",
                                  "Largest Win", "Largest Loss") else 2
            pfx = "$" if label not in ("Win Rate (%)", "Profit Factor", "Total Trades",
                                        "Winning Trades", "Losing Trades",
                                        "Max Consec. Wins", "Max Consec. Losses",
                                        "Average Trade") else ""
            sfx = "%" if label == "Win Rate (%)" else ""
            return [label, _fmt(fv, dec, pfx, sfx)]
        except Exception:
            return [label, str(val) if val is not None else "—"]

    l_rows = [["Metric", "Value"]] + [_kpm_row(l, v) for l, v in left_kpm]
    r_rows = [["Metric", "Value"]] + [_kpm_row(l, v) for l, v in right_kpm]
    while len(r_rows) < len(l_rows):
        r_rows.append(["", ""])

    combo_rows = [l + [""] + r for l, r in zip(l_rows, r_rows)]
    cw2 = [cw * 0.27, cw * 0.19, cw * 0.08, cw * 0.27, cw * 0.19]
    combo_style = [
        ("BACKGROUND",     (0, 0), (1,  0), CARD_BG),
        ("BACKGROUND",     (3, 0), (4,  0), CARD_BG),
        ("TEXTCOLOR",      (0, 0), (-1,-1), TEXT_PRIMARY),
        ("FONTNAME",       (0, 0), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0, 1), (1, -1), [ROW_ODD, ROW_EVEN]),
        ("ROWBACKGROUNDS", (3, 1), (4, -1), [ROW_ODD, ROW_EVEN]),
        ("GRID",           (0, 0), (1, -1), 0.3, BORDER),
        ("GRID",           (3, 0), (4, -1), 0.3, BORDER),
        ("FONTNAME",       (0, 0), (0,  0), "Helvetica-Bold"),
        ("FONTNAME",       (3, 0), (3,  0), "Helvetica-Bold"),
        ("TEXTCOLOR",      (0, 0), (1,  0), TEXT_MUTED),
        ("TEXTCOLOR",      (3, 0), (4,  0), TEXT_MUTED),
        ("ALIGN",          (1, 0), (1, -1), "RIGHT"),
        ("ALIGN",          (4, 0), (4, -1), "RIGHT"),
        ("LEFTPADDING",    (0, 0), (-1,-1), 6),
        ("RIGHTPADDING",   (0, 0), (-1,-1), 6),
        ("TOPPADDING",     (0, 0), (-1,-1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1,-1), 3),
    ]
    for i, row in enumerate(combo_rows[1:], 1):
        for col_idx, met_idx in [(1, 0), (4, 3)]:
            try:
                fv = float(str(row[col_idx]).replace("$","").replace("%","").replace(",",""))
                if any(k in str(row[met_idx]) for k in ("Profit","Win","Loss","Net","Avg")):
                    clr = ACCENT_GREEN if fv >= 0 else ACCENT_RED
                    combo_style.append(("TEXTCOLOR", (col_idx, i), (col_idx, i), clr))
            except Exception:
                pass
    combo_tbl = Table(combo_rows, colWidths=cw2)
    combo_tbl.setStyle(TableStyle(combo_style))
    story.append(combo_tbl)
    story.append(Spacer(1, 6 * mm))

    # ── Equity chart ──────────────────────────────────────────────────────────
    story.append(Paragraph("Daily Equity Curve", section_s))
    eq = _chart_equity(data, cw, 200)
    if eq:
        story.append(eq)
    story.append(Spacer(1, 6 * mm))

    # ── Time-based analysis ───────────────────────────────────────────────────
    story.append(Paragraph("Time-Based Performance Analysis", section_s))
    hw = (cw - 6 * mm) / 2

    dow_tbl = _make_styled_table(
        [["Day", "Total P&L", "Trades", "Avg P&L", "Win Rate %"]] +
        [[r[0], _fmt(r[1], 2, "$"), _fmt(r[2], 0), _fmt(r[3], 2, "$"), _fmt(r[4], 2, suffix="%")]
         for r in data["dow"]],
        [hw * f for f in [0.28, 0.22, 0.18, 0.18, 0.14]], pnl_col=1)

    hour_tbl = _make_styled_table(
        [["Hour", "Total P&L", "Trades", "Avg P&L", "Win Rate %"]] +
        [[f"{int(r[0]):02d}:00" if r[0] is not None else "?",
          _fmt(r[1], 2, "$"), _fmt(r[2], 0), _fmt(r[3], 2, "$"), _fmt(r[4], 2, suffix="%")]
         for r in data["hour"]],
        [hw * f for f in [0.22, 0.22, 0.18, 0.20, 0.18]], pnl_col=1)

    def _side(a, b):
        t = Table([[a, "", b]], colWidths=[hw, 6 * mm, hw])
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                ("LEFTPADDING",  (0, 0), (-1, -1), 0),
                                ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
        return t

    story.append(_side(dow_tbl, hour_tbl))
    story.append(Spacer(1, 4 * mm))

    dc = _chart_dow(data["dow"],  hw, 150)
    hc = _chart_hour(data["hour"], hw, 150)
    if dc and hc:
        story.append(_side(dc, hc))
    story.append(Spacer(1, 6 * mm))

    # ── Monthly ───────────────────────────────────────────────────────────────
    story.append(Paragraph("Monthly Performance", section_s))
    story.append(_make_styled_table(
        [["Month", "Total P&L", "Trades", "Avg P&L", "Win Rate %"]] +
        [[r[0], _fmt(r[1], 2, "$"), _fmt(r[2], 0), _fmt(r[3], 2, "$"), _fmt(r[4], 2, suffix="%")]
         for r in data["month"]],
        [cw * f for f in [0.22, 0.20, 0.16, 0.20, 0.22]], pnl_col=1))
    story.append(Spacer(1, 6 * mm))

    # ── Distribution + Risk ───────────────────────────────────────────────────
    story.append(Paragraph("Distribution & Risk Metrics", section_s))

    dist_tbl = _make_styled_table(
        [["P&L Range", "Count", "%"]] +
        [[r[0], _fmt(r[1], 0), _fmt(r[2], 2, suffix="%")] for r in data["dist"]],
        [hw * f for f in [0.55, 0.22, 0.23]])

    risk_fmt = {
        "Max Drawdown ($)":     lambda v: _fmt(v, 2, "$"),
        "Max Drawdown (%)":     lambda v: _fmt(v, 2, suffix="%"),
        "Sharpe-like Ratio":    lambda v: _fmt(v, 4),
        "Expectancy per Trade": lambda v: _fmt(v, 4),
        "Std Dev of Returns":   lambda v: _fmt(v, 4),
        "Avg Return per Trade": lambda v: _fmt(v, 4),
    }
    risk_tbl = _make_styled_table(
        [["Risk Metric", "Value"]] +
        [[lbl, risk_fmt.get(lbl, lambda v: _fmt(v, 4))(val)]
         for lbl, val in data["risk"].items()],
        [hw * f for f in [0.62, 0.38]])

    story.append(_side(dist_tbl, risk_tbl))
    story.append(Spacer(1, 4 * mm))

    distc = _chart_dist(data["dist"], cw * 0.6, 130)
    if distc:
        story.append(distc)
    story.append(Spacer(1, 6 * mm))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER,
                             spaceBefore=4, spaceAfter=4))
    story.append(Paragraph(
        f"<font color='#8B949E' size='7'>Report generated from TradingView backtest data "
        f"&bull; {data['stock']} &bull; {data['date_range']} &bull; "
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
        ps("F", alignment=TA_CENTER)))

    def _bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(DARK_BG)
        canvas.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_bg, onLaterPages=_bg)
    print(f"  ✓ PDF saved: {output_path}")


# =============================================================================
# SECTION 3 — PROCESS FILE (ties everything together)
# =============================================================================

def process_file(input_file, output_file=None):
    """Process a single TradingView backtest file → Excel + PDF."""
    if not os.path.exists(input_file):
        print(f"Error: File not found: {input_file}")
        return False

    print(f"Processing: {input_file}")

    try:
        trades_df = pd.read_excel(input_file, sheet_name='List of trades')
        print(f"  Loaded {len(trades_df)} trade records")

        metrics, data_frames = analyze_trades(trades_df)
        if metrics is None:
            print("  Error: No valid trade data found")
            return False

        print(f"  Analyzed {metrics['total_trades']} trades  |  "
              f"Win Rate: {metrics['win_rate']:.2f}%  |  "
              f"Profit Factor: {metrics['profit_factor']:.2f}")

        wb = load_workbook(input_file)
        create_analysis_sheet(wb, metrics, data_frames)

        # Determine output paths
        if output_file is None:
            base, ext = os.path.splitext(input_file)
            output_file = f"{base}_analyzed{ext}"

        pdf_file = os.path.splitext(output_file)[0] + ".pdf"

        # Save Excel
        wb.save(output_file)
        print(f"  ✓ Excel saved: {output_file}")

        # Generate PDF directly from in-memory data (no re-read needed)
        pdf_data = build_pdf_data(input_file, wb, metrics, data_frames)
        build_pdf(pdf_data, pdf_file)

        return True

    except Exception as e:
        print(f"  Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def batch_process(directory):
    """Process all Excel files in a directory → Excel + PDF for each."""
    if not os.path.isdir(directory):
        print(f"Error: Directory not found: {directory}")
        return

    files = [f for f in os.listdir(directory)
             if f.endswith('.xlsx') and not f.startswith('~') and '_analyzed' not in f]

    if not files:
        print(f"No Excel files found in {directory}")
        return

    print(f"Found {len(files)} file(s) to process\n")
    success = 0
    for i, filename in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {filename}")
        if process_file(os.path.join(directory, filename)):
            success += 1
        print()

    print(f"Batch complete: {success}/{len(files)} files processed successfully")


# =============================================================================
# ENTRY POINT
# =============================================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    if sys.argv[1] == '--batch':
        if len(sys.argv) < 3:
            print("Error: Please specify a directory for batch processing")
            sys.exit(1)
        batch_process(sys.argv[2])
    else:
        input_file  = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None

        if process_file(input_file, output_file):
            print("\n✓ Analysis complete! — Excel + PDF generated.")
        else:
            print("\n✗ Analysis failed")
            sys.exit(1)


if __name__ == '__main__':
    main()
